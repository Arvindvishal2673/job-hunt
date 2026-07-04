"""Styled Excel writer with conditional color highlights (openpyxl)."""

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import CandidateProfile, JobListing

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
DECISION_FILLS = {
    "Strong Fit": PatternFill("solid", fgColor="C6EFCE"),  # green
    "Decent Fit": PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "Weak Fit": PatternFill("solid", fgColor="FFC7CE"),  # red
}
COLUMNS = [
    "Fit Score", "Fit Decision", "Title", "Company", "Location",
    "Source", "Salary", "Match Reasons", "Gaps Identified", "URL",
]
WIDTHS = [10, 14, 40, 24, 22, 16, 18, 50, 50, 60]
WRAP_COLUMNS = (8, 9)


def write_excel(
    profile: CandidateProfile,
    jobs: List[JobListing],
    output_path: str = "outputs/job_matches.xlsx",
) -> str:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Job Matches"

    for col, (name, width) in enumerate(zip(COLUMNS, WIDTHS), start=1):
        cell = sheet.cell(row=1, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A2"

    for row, job in enumerate(jobs, start=2):
        values = [
            round(job.fit_score, 1),
            job.fit_decision or "Unrated",
            job.title,
            job.company,
            job.location,
            job.source,
            job.salary,
            "\n".join(f"\u2022 {r}" for r in job.fit_reasons),
            "\n".join(f"\u2022 {g}" for g in job.gaps_identified),
            job.url,
        ]
        for col, value in enumerate(values, start=1):
            cell = sheet.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=col in WRAP_COLUMNS)
        decision_fill = DECISION_FILLS.get(job.fit_decision)
        if decision_fill:
            sheet.cell(row=row, column=2).fill = decision_fill
        if job.url:
            link = sheet.cell(row=row, column=10)
            link.hyperlink = job.url
            link.font = Font(color="0563C1", underline="single")

    summary = workbook.create_sheet("Candidate Profile")
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 90
    profile_rows = [
        ("Summary", profile.summary),
        ("Seniority", profile.seniority),
        ("Skills", ", ".join(profile.skills)),
        ("Target Titles", ", ".join(profile.job_titles)),
        ("Search Queries", "; ".join(profile.search_queries)),
    ]
    for row, (key, value) in enumerate(profile_rows, start=1):
        summary.cell(row=row, column=1, value=key).font = Font(bold=True)
        summary.cell(row=row, column=2, value=value).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    workbook.save(path)
    return str(path)
